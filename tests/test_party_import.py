from __future__ import annotations

import io

from openpyxl import Workbook

from app.extensions import db
from app.models.models import Party, PartyType

from conftest import login


def _xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.append(["Ad Soyad", "Telefon", "E-posta", "Adres", "Vergi No", "Not"])
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def test_import_parties_adds_new_dentists(client, app):
    login(client, "admin", "admin-pass")
    data = {
        "file": (
            _xlsx([
                ["Dr. Excel Bir", "+905551112233", "bir@ornek.com", "İstanbul", "1234567890", "VIP"],
                ["Dr. Excel İki", None, None, None, None, None],
                [None, "+905550000000", None, None, None, None],  # adsız satır atlanır
            ]),
            "doktorlar.xlsx",
        )
    }
    response = client.post(
        "/parties/import", data=data, content_type="multipart/form-data", follow_redirects=True
    )
    assert response.status_code == 200
    text = response.get_data(as_text=True)
    assert "2 yeni" in text
    assert "1 atlandı" in text

    with app.app_context():
        party = db.session.execute(
            db.select(Party).where(Party.name == "Dr. Excel Bir")
        ).scalar_one()
        assert party.party_type == PartyType.DENTIST
        assert party.phone == "+905551112233"
        assert party.email == "bir@ornek.com"
        assert party.tax_id == "1234567890"
        assert party.is_active is True


def test_import_parties_updates_existing_by_name(client, app):
    login(client, "admin", "admin-pass")
    with app.app_context():
        db.session.add(Party(party_type=PartyType.DENTIST, name="Dr. Mevcut", phone="+905551110001"))
        db.session.commit()

    data = {"file": (_xlsx([["dr. mevcut", "+905559998877", "yeni@ornek.com", None, None, None]]), "d.xlsx")}
    response = client.post(
        "/parties/import", data=data, content_type="multipart/form-data", follow_redirects=True
    )
    text = response.get_data(as_text=True)
    assert "1 güncellendi" in text

    with app.app_context():
        parties = db.session.execute(
            db.select(Party).where(db.func.lower(Party.name) == "dr. mevcut")
        ).scalars().all()
        assert len(parties) == 1
        assert parties[0].phone == "+905559998877"
        assert parties[0].email == "yeni@ornek.com"


def test_import_parties_rejects_non_excel(client, app):
    login(client, "admin", "admin-pass")
    data = {"file": (io.BytesIO(b"not excel"), "liste.csv")}
    response = client.post(
        "/parties/import", data=data, content_type="multipart/form-data", follow_redirects=True
    )
    assert "Yalnızca .xlsx veya .xls" in response.get_data(as_text=True)


def test_import_parties_corrupt_file_shows_friendly_message(client, app):
    """A .xlsx-named file that isn't actually a valid zip/xlsx must not leak
    the raw exception text to the user (previously: f"...: {str(e)}")."""
    login(client, "admin", "admin-pass")
    data = {"file": (io.BytesIO(b"this is not a real xlsx file"), "bozuk.xlsx")}
    response = client.post(
        "/parties/import", data=data, content_type="multipart/form-data", follow_redirects=True
    )
    assert response.status_code == 200
    text = response.get_data(as_text=True)
    assert "Dosya okunamadı" in text
    assert "BadZipFile" not in text
    assert "Traceback" not in text


def test_import_parties_neutralizes_formula_prefix_in_notes(client, app):
    """A note starting with '=' must be defused before it's stored, so a
    later Excel/CSV export of this data can't auto-execute a formula for
    whoever opens it.

    Written as a literal text cell (data_type='s') rather than via plain
    assignment: openpyxl (like Excel) treats a normally-assigned string
    starting with '=' as an actual formula, which load_workbook(data_only=
    True) then reads back as None (no cached result) instead of the raw
    text — that would test nothing. A cell explicitly typed as text is
    what a real attacker-crafted "note" column looks like.
    """
    login(client, "admin", "admin-pass")
    payload = "=cmd|'/c calc'!A1"

    wb = Workbook()
    ws = wb.active
    ws.append(["Ad Soyad", "Telefon", "E-posta", "Adres", "Vergi No", "Not"])
    ws.append(["Dr. Formula Test", "+905551119999", None, None, None, None])
    notes_cell = ws.cell(row=2, column=6)
    notes_cell.value = payload
    notes_cell.data_type = "s"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    data = {"file": (buffer, "formula.xlsx")}
    response = client.post(
        "/parties/import", data=data, content_type="multipart/form-data", follow_redirects=True
    )
    assert response.status_code == 200

    with app.app_context():
        party = db.session.execute(
            db.select(Party).where(Party.name == "Dr. Formula Test")
        ).scalar_one()
        assert party.notes == "'" + payload
        assert not party.notes.startswith("=")
