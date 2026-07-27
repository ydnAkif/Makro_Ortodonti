import logging
from zipfile import BadZipFile

from flask import Blueprint, render_template, redirect, url_for, flash, request, Response
from flask_login import login_required
from datetime import date, timedelta
from decimal import Decimal

from app.extensions import db
from app.models.models import Party, PartyType, WorkOrder, Treatment, TreatmentCategory, ExchangeRate, Makbuz, MakbuzPayment, money
from app.authz import permissions_required
from app.services.party_service import PartyService
from app.services.search_service import tr_contains, tr_order, tr_fold, tr_equals
from app.services.validation_service import normalize_display_name, neutralize_formula_prefix, parse_date
from app.services.exchange_service import get_latest_usd_rate
from app.services.reports_pdf_service import generate_work_orders_pdf
from app.services.settings_service import get_clinic_identity
from app.constants import MONTHS

logger = logging.getLogger(__name__)

parties_bp = Blueprint("parties", __name__)


def _get_treatments_by_category(category):

    treatments = db.session.execute(
        db.select(Treatment)
        .where(Treatment.category == category, Treatment.is_active.is_(True))
        .order_by(tr_order(Treatment.name))
    ).scalars().all()
    return [{"id": t.id, "name": t.name, "price": float(t.price_eur), "currency": t.currency} for t in treatments]


def _get_current_rate():
    rate = db.session.execute(
        db.select(ExchangeRate).order_by(ExchangeRate.rate_date.desc()).limit(1)
    ).scalar_one_or_none()
    return float(rate.eur_to_try) if rate else 0


@parties_bp.route("/")
@login_required
@permissions_required("clinical.view")
def list_parties():
    search = request.args.get("search", "").strip()
    today = date.today()

    query = db.select(Party).where(
        Party.party_type == PartyType.DENTIST,
        Party.is_active.is_(True),
    )

    if search:
        phone_digits = search.replace(" ", "").replace("-", "")
        query = query.where(
            db.or_(
                tr_contains(Party.name, search),
                tr_contains(Party.tax_id, search),
                tr_contains(Party.email, search),
                db.func.replace(db.func.replace(Party.phone, " ", ""), "-", "").ilike(
                    f"%{phone_digits}%"
                ),
            )
        )

    query = query.order_by(tr_order(Party.name))
    pagination = db.paginate(query, page=max(request.args.get("page", 1, type=int), 1), per_page=25, max_per_page=100, error_out=False)
    parties = pagination.items

    page_party_ids = [p.id for p in parties]
    kasa_by_party: dict[int, dict] = {}

    if page_party_ids:
        all_makbuzlar = db.session.execute(
            db.select(Makbuz).where(Makbuz.party_id.in_(page_party_ids))
        ).scalars().all()

        makbuzlar_by_party: dict[int, list[Makbuz]] = {}
        for m in all_makbuzlar:
            makbuzlar_by_party.setdefault(m.party_id, []).append(m)

        for p in parties:
            party_makbuzlar = makbuzlar_by_party.get(p.id, [])
            billed = money(sum((m.grand_total for m in party_makbuzlar if m.status in (Makbuz.STATUS_SENT, Makbuz.STATUS_PAID)), Decimal("0.00")))
            paid = money(sum((m.collected_amount for m in party_makbuzlar), Decimal("0.00")))
            prev_bal = money(p.previous_balance or Decimal("0.00"))
            total_balance = money(billed - paid + prev_bal)

            current_m = next((m for m in party_makbuzlar if m.year == today.year and m.month == today.month), None)

            kasa_by_party[p.id] = {
                "work_total": money(current_m.grand_total) if current_m else Decimal("0.00"),
                "vat_total": (money(current_m.vat_amount) if current_m.vat_applied else None) if current_m else None,
                "payment_total": money(current_m.collected_amount) if current_m else Decimal("0.00"),
                "balance": total_balance,
            }
    else:
        for p in parties:
            kasa_by_party[p.id] = {
                "work_total": Decimal("0.00"),
                "vat_total": None,
                "payment_total": Decimal("0.00"),
                "balance": money(p.previous_balance or Decimal("0.00")),
            }

    # System-wide overall metrics for top summary cards
    total_doctors_count = db.session.scalar(
        db.select(db.func.count(Party.id)).where(Party.party_type == PartyType.DENTIST, Party.is_active.is_(True))
    ) or 0

    current_month_start = date(today.year, today.month, 1)
    next_month_start = date(today.year + (today.month == 12), today.month % 12 + 1, 1)

    current_month_work_total = db.session.scalar(
        db.select(db.func.coalesce(db.func.sum(WorkOrder.total_price), Decimal("0.00")))
        .where(
            WorkOrder.work_date >= current_month_start,
            WorkOrder.work_date < next_month_start,
        )
    ) or Decimal("0.00")

    all_active_parties = db.session.execute(
        db.select(Party.id, Party.previous_balance).where(Party.party_type == PartyType.DENTIST, Party.is_active.is_(True))
    ).all()
    all_party_ids = [p.id for p in all_active_parties]
    sum_prev_balances = money(sum((p.previous_balance or Decimal("0.00") for p in all_active_parties), Decimal("0.00")))

    if all_party_ids:
        total_billed = db.session.scalar(
            db.select(db.func.coalesce(db.func.sum(Makbuz.grand_total), Decimal("0.00")))
            .where(
                Makbuz.party_id.in_(all_party_ids),
                Makbuz.status.in_((Makbuz.STATUS_SENT, Makbuz.STATUS_PAID))
            )
        ) or Decimal("0.00")

        total_payments_entries = db.session.scalar(
            db.select(db.func.coalesce(db.func.sum(MakbuzPayment.amount), Decimal("0.00")))
            .join(Makbuz, MakbuzPayment.makbuz_id == Makbuz.id)
            .where(Makbuz.party_id.in_(all_party_ids))
        ) or Decimal("0.00")

        total_legacy_paid = db.session.scalar(
            db.select(db.func.coalesce(db.func.sum(Makbuz.paid_amount), Decimal("0.00")))
            .where(
                Makbuz.party_id.in_(all_party_ids),
                ~Makbuz.payment_entries.any(),
            )
        ) or Decimal("0.00")

        total_collected = money(total_payments_entries + total_legacy_paid)
    else:
        total_billed = Decimal("0.00")
        total_collected = Decimal("0.00")

    total_clinic_balance = money(total_billed - total_collected + sum_prev_balances)

    kasa_page_total_work = money(sum((row["work_total"] for row in kasa_by_party.values()), Decimal("0.00")))
    kasa_page_total_payment = money(sum((row["payment_total"] for row in kasa_by_party.values()), Decimal("0.00")))
    kasa_page_total_balance = money(sum((row["balance"] for row in kasa_by_party.values()), Decimal("0.00")))

    # Canlı arama yalnızca sonuç tablosunu ister; sayfa yeniden yüklenmez.
    template = (
        "parties/_results.html"
        if request.args.get("partial") == "1"
        else "parties/list.html"
    )
    return render_template(
        template,
        parties=parties,
        search=search,
        pagination=pagination,
        summary_month=today,
        kasa_by_party=kasa_by_party,
        kasa_page_total_work=kasa_page_total_work,
        kasa_page_total_payment=kasa_page_total_payment,
        kasa_page_total_balance=kasa_page_total_balance,
        total_doctors_count=total_doctors_count,
        current_month_work_total=current_month_work_total,
        total_clinic_balance=total_clinic_balance,
    )


def _resolve_work_order_query(request_args):
    """Helper to parse work order list filters and return period data & query results."""
    today = date.today()
    view = request_args.get("view", "month")
    if view not in ("day", "month", "year", "all"):
        view = "month"

    search = request_args.get("search", "").strip()

    year = request_args.get("year", today.year, type=int)
    month = request_args.get("month", today.month, type=int)
    date_str = request_args.get("date", "").strip()
    selected_date = parse_date(date_str) or today

    previous_date = (selected_date - timedelta(days=1)).isoformat()
    next_date = (selected_date + timedelta(days=1)).isoformat()

    prev_nav_month = 12 if month == 1 else month - 1
    prev_nav_year = year - 1 if month == 1 else year
    next_nav_month = 1 if month == 12 else month + 1
    next_nav_year = year + 1 if month == 12 else year

    if view == "all":
        period_start = date(2000, 1, 1)
        period_end = date(2099, 12, 31)
        period_label = "Tüm Zamanlar"
    elif view == "year":
        period_start = date(year, 1, 1)
        period_end = date(year + 1, 1, 1)
        period_label = f"{year} Yılı"
    elif view == "day":
        period_start = selected_date
        period_end = selected_date + timedelta(days=1)
        period_label = selected_date.strftime("%d.%m.%Y")
    else:  # view == "month"
        try:
            period_start = date(year, month, 1)
        except ValueError:
            year, month = today.year, today.month
            period_start = date(year, month, 1)
        period_end = date(year + (month == 12), month % 12 + 1, 1)
        period_label = f"{MONTHS[month - 1][1]} {year}"

    query = (
        db.select(WorkOrder)
        .join(Party, WorkOrder.party_id == Party.id)
        .where(
            WorkOrder.work_date >= period_start,
            WorkOrder.work_date < period_end,
            Party.is_active.is_(True),
        )
    )
    if search:
        query = query.where(db.or_(
            tr_contains(Party.name, search),
            tr_contains(WorkOrder.patient_name, search),
            tr_contains(WorkOrder.apparatus_type, search),
            tr_contains(WorkOrder.extra_addons, search),
            tr_contains(WorkOrder.notes, search),
        ))

    work_orders = db.session.execute(
        query.order_by(WorkOrder.work_date.desc(), WorkOrder.id.desc())
    ).scalars().all()

    return {
        "work_orders": work_orders,
        "view": view,
        "selected_date": selected_date,
        "year": year,
        "month": month,
        "months": MONTHS,
        "years": list(range(today.year - 3, today.year + 2)),
        "period_label": period_label,
        "search": search,
        "doctor_count": len({wo.party_id for wo in work_orders}),
        "period_total": money(sum((wo.total_price for wo in work_orders), Decimal("0.00"))),
        "today": today,
        "previous_date": previous_date,
        "next_date": next_date,
        "previous_month": prev_nav_month,
        "previous_year": prev_nav_year,
        "next_month": next_nav_month,
        "next_year": next_nav_year,
    }


@parties_bp.route("/work-orders", strict_slashes=False)
@login_required
@permissions_required("clinical.view")
def list_work_orders():
    """Operational ledger for all doctors' work orders with daily/monthly/yearly/all views."""
    data = _resolve_work_order_query(request.args)
    return render_template("work_orders/list.html", **data)


@parties_bp.route("/work-orders/pdf", strict_slashes=False)
@login_required
@permissions_required("clinical.view")
def export_work_orders_pdf():
    """PDF export for filtered work orders."""
    data = _resolve_work_order_query(request.args)
    clinic = get_clinic_identity()

    pdf_bytes = generate_work_orders_pdf(
        clinic_name=clinic["clinic_name"],
        clinic_phone=clinic["clinic_phone"],
        clinic_email=clinic["clinic_email"],
        period_label=data["period_label"],
        work_orders=data["work_orders"],
        doctor_count=data["doctor_count"],
        period_total=data["period_total"],
    )

    filename = f"is_emirleri_{data['view']}_{data['period_label'].replace(' ', '_').lower()}.pdf"
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@parties_bp.route("/<int:party_id>/summary/pdf", strict_slashes=False)
@login_required
@permissions_required("clinical.view")
def export_party_summary_pdf(party_id):
    """PDF export for a single doctor's monthly or yearly work-order summary."""
    party = db.get_or_404(Party, party_id)
    today = date.today()

    view = request.args.get("view", "month")
    year = request.args.get("year", today.year, type=int)
    month = request.args.get("month", today.month, type=int)

    if view == "year":
        period_start = date(year, 1, 1)
        period_end = date(year + 1, 1, 1)
        period_label = f"{party.display_name} · {year} Yılı"
    else:
        view = "month"
        try:
            period_start = date(year, month, 1)
        except ValueError:
            year, month = today.year, today.month
            period_start = date(year, month, 1)
        period_end = date(year + (month == 12), month % 12 + 1, 1)
        period_label = f"{party.display_name} · {MONTHS[month - 1][1]} {year}"

    work_orders = db.session.execute(
        db.select(WorkOrder)
        .where(
            WorkOrder.party_id == party.id,
            WorkOrder.work_date >= period_start,
            WorkOrder.work_date < period_end,
        )
        .order_by(WorkOrder.work_date.desc(), WorkOrder.id.desc())
    ).scalars().all()

    clinic = get_clinic_identity()

    period_total = money(sum((wo.total_price for wo in work_orders), Decimal("0.00")))
    pdf_bytes = generate_work_orders_pdf(
        clinic_name=clinic["clinic_name"],
        clinic_phone=clinic["clinic_phone"],
        clinic_email=clinic["clinic_email"],
        period_label=period_label,
        work_orders=work_orders,
        doctor_count=1,
        period_total=period_total,
    )

    safe_name = "_".join((party.display_name or "doktor").lower().split())
    filename = f"ozet_{safe_name}_{view}_{year}{f'_{month:02d}' if view == 'month' else ''}.pdf"
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@parties_bp.route("/add", methods=["GET", "POST"])
@login_required
@permissions_required("clinical.edit")
def add_party():
    if request.method == "POST":
        party = Party(
            party_type=PartyType.DENTIST,
            name=normalize_display_name(request.form.get("name", "")),
            phone=request.form.get("phone", "").strip() or None,
            email=request.form.get("email", "").strip() or None,
            address=request.form.get("address", "").strip() or None,
            tax_id=request.form.get("tax_id", "").strip() or None,
            notes=request.form.get("notes", "").strip() or None,
            is_active=request.form.get("is_active") == "on",
            previous_balance=money(request.form.get("previous_balance", "0") or "0"),
        )
        db.session.add(party)
        db.session.commit()
        flash(f"{party.display_name} başarıyla eklendi.", "success")
        return redirect(url_for("parties.detail_party", party_id=party.id))

    return render_template("parties/form.html", party=None)


@parties_bp.route("/import", methods=["GET", "POST"])
@login_required
@permissions_required("clinical.edit")
def import_parties():
    """Diş hekimlerini Excel dosyasından toplu içe aktar (A: Ad, B: Telefon, C: E-posta, D: Adres, E: Vergi/TC No, F: Not)."""
    if request.method == "POST":
        file = request.files.get("file")
        if not file:
            flash("Dosya seçilmedi.", "danger")
            return redirect(url_for("parties.import_parties"))

        filename = (file.filename or "").lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
            flash("Yalnızca .xlsx veya .xls dosyaları desteklenir.", "danger")
            return redirect(url_for("parties.import_parties"))

        try:
            import io
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
            ws = wb.active
            added, updated, skipped = 0, 0, 0

            def _text(value, limit):
                return str(value).strip()[:limit] if value is not None and str(value).strip() else None

            for row in ws.iter_rows(min_row=2, values_only=True):
                name = _text(row[0] if row else None, 200)
                if not name:
                    skipped += 1
                    continue
                name = normalize_display_name(neutralize_formula_prefix(name))
                phone = _text(row[1] if len(row) > 1 else None, 20)
                email = _text(row[2] if len(row) > 2 else None, 200)
                address = neutralize_formula_prefix(_text(row[3] if len(row) > 3 else None, 2000))
                tax_id = _text(row[4] if len(row) > 4 else None, 50)
                notes = neutralize_formula_prefix(_text(row[5] if len(row) > 5 else None, 2000))

                # Yalnızca isimle eşleştir (Türkçe duyarlı): aynı telefonu paylaşan
                # farklı isimli kayıtlar (ör. iki şubeli klinik) ayrı satır kalmalı.
                existing = db.session.execute(
                    db.select(Party).where(
                        Party.party_type == PartyType.DENTIST,
                        tr_equals(Party.name, name),
                    )
                ).scalars().first()

                if existing:
                    existing.name = name
                    existing.phone = phone or existing.phone
                    existing.email = email or existing.email
                    existing.address = address or existing.address
                    existing.tax_id = tax_id or existing.tax_id
                    existing.notes = notes or existing.notes
                    existing.is_active = True
                    updated += 1
                else:
                    db.session.add(Party(
                        party_type=PartyType.DENTIST, name=name, phone=phone,
                        email=email, address=address, tax_id=tax_id, notes=notes,
                        is_active=True,
                    ))
                    added += 1

            db.session.commit()
            wb.close()
            flash(f"İçe aktarma tamamlandı: {added} yeni, {updated} güncellendi, {skipped} atlandı.", "success")
            return redirect(url_for("parties.list_parties"))
        except (KeyError, IndexError, ValueError, OSError, BadZipFile) as exc:
            db.session.rollback()
            logger.warning("Hekim içe aktarma: dosya okunamadı: %s", exc)
            flash("Dosya okunamadı. Geçerli bir .xlsx dosyası olduğundan ve beklenen sütun düzenine "
                  "sahip olduğundan emin olun.", "danger")
            return redirect(url_for("parties.import_parties"))
        except Exception:
            db.session.rollback()
            logger.exception("Hekim içe aktarma sırasında beklenmeyen hata")
            flash("İçe aktarma sırasında beklenmeyen bir hata oluştu.", "danger")
            return redirect(url_for("parties.import_parties"))

    return render_template("parties/import.html")


@parties_bp.route("/<int:party_id>")
@login_required
@permissions_required("clinical.view")
def detail_party(party_id):
    party = db.get_or_404(Party, party_id)

    today = date.today()
    view = request.args.get("view", "month")
    year = request.args.get("year", today.year, type=int)

    if view == "year":
        month = None
        period_start = date(year, 1, 1)
        period_end = date(year + 1, 1, 1)
    else:
        view = "month"
        month = request.args.get("month", today.month, type=int)
        try:
            period_start = date(year, month, 1)
        except ValueError:
            year, month = today.year, today.month
            period_start = date(year, month, 1)
        period_end = date(year + (month == 12), month % 12 + 1, 1)

    period_work_orders = db.session.execute(
        db.select(WorkOrder)
        .where(
            WorkOrder.party_id == party_id,
            WorkOrder.work_date >= period_start,
            WorkOrder.work_date < period_end,
        )
        .order_by(WorkOrder.work_date.desc(), WorkOrder.id.desc())
    ).scalars().all()

    search = request.args.get("search", "").strip()
    if search:
        folded_search = tr_fold(search)
        work_orders = [
            wo for wo in period_work_orders
            if folded_search in tr_fold(" ".join(filter(None, [
                wo.patient_name, wo.apparatus_type, wo.extra_addons, wo.notes,
            ])))
        ]
    else:
        work_orders = period_work_orders

    total_apparatus = sum((wo.apparatus_price for wo in period_work_orders), Decimal("0.00"))
    total_extra = sum((wo.extra_price for wo in period_work_orders), Decimal("0.00"))
    total_overall = sum((wo.total_price for wo in period_work_orders), Decimal("0.00"))

    period_makbuz = None
    if view == "month":
        period_makbuz = db.session.execute(
            db.select(Makbuz).where(
                Makbuz.party_id == party_id,
                Makbuz.year == year,
                Makbuz.month == month,
            )
        ).scalar_one_or_none()
    total_vat = period_makbuz.vat_amount if period_makbuz else Decimal("0.00")
    total_with_vat = total_overall + total_vat

    year_start = date(year, 1, 1)
    year_end = date(year + 1, 1, 1)
    year_work_orders = db.session.execute(
        db.select(WorkOrder)
        .where(
            WorkOrder.party_id == party_id,
            WorkOrder.work_date >= year_start,
            WorkOrder.work_date < year_end,
        )
    ).scalars().all()
    year_total_apparatus = sum((wo.apparatus_price for wo in year_work_orders), Decimal("0.00"))
    year_total_extra = sum((wo.extra_price for wo in year_work_orders), Decimal("0.00"))
    year_total_overall = sum((wo.total_price for wo in year_work_orders), Decimal("0.00"))
    year_work_order_count = len(year_work_orders)

    monthly_totals = _compute_monthly_totals(party_id, year, year_work_orders)

    previous_periods_debt = _compute_previous_debt(party_id, year, month, view)

    devreden_total = previous_periods_debt + party.previous_balance

    oldest_work_date = db.session.scalar(
        db.select(db.func.min(WorkOrder.work_date)).where(WorkOrder.party_id == party_id)
    )
    first_year = min(oldest_work_date.year if oldest_work_date else today.year, year)
    last_year = max(today.year + 1, year)

    if view == "month":
        previous_month_date = period_start - timedelta(days=1)
        prev_nav_year = previous_month_date.year
        prev_nav_month = previous_month_date.month
        next_nav_year = period_end.year
        next_nav_month = period_end.month
    else:
        prev_nav_year = year - 1
        prev_nav_month = 1
        next_nav_year = year + 1
        next_nav_month = 1

    if view == "month":
        period_label = f"{MONTHS[month - 1][1]} {year}"
    else:
        period_label = f"{year} Yılı"

    return render_template(
        "parties/detail.html",
        party=party,
        work_orders=work_orders,
        period_work_order_count=len(period_work_orders),
        search=search,
        view=view,
        year=year,
        month=month,
        months=MONTHS,
        years=range(last_year, first_year - 1, -1),
        period_label=period_label,
        previous_year=prev_nav_year,
        previous_month=prev_nav_month,
        next_year=next_nav_year,
        next_month=next_nav_month,
        total_apparatus=total_apparatus,
        total_extra=total_extra,
        total_overall=total_overall,
        total_vat=total_vat,
        total_with_vat=total_with_vat,
        year_total_apparatus=year_total_apparatus,
        year_total_extra=year_total_extra,
        year_total_overall=year_total_overall,
        year_work_order_count=year_work_order_count,
        monthly_totals=monthly_totals,
        previous_periods_debt=previous_periods_debt,
        devreden_total=devreden_total,
        period_makbuz=period_makbuz,
        period_makbuz_status=(
            {
                Makbuz.STATUS_DRAFT: "Taslak",
                Makbuz.STATUS_SENT: "Gönderildi",
                Makbuz.STATUS_PAID: "Ödendi",
            }.get(period_makbuz.status, period_makbuz.status)
            if period_makbuz else None
        ),
    )


def _get_party_outstanding(party: Party) -> Decimal:
    """Doktorun devreden borç ve ödenmemiş makbuzlarının toplam açık bakiyesini hesapla."""
    makbuzlar = db.session.execute(
        db.select(Makbuz).where(Makbuz.party_id == party.id)
    ).scalars().all()

    billed = money(sum((m.grand_total for m in makbuzlar if m.status in (Makbuz.STATUS_SENT, Makbuz.STATUS_PAID)), Decimal("0.00")))
    paid = money(sum((m.collected_amount for m in makbuzlar), Decimal("0.00")))
    prev_bal = money(party.previous_balance or Decimal("0.00"))
    return money(billed - paid + prev_bal)


def _compute_monthly_totals(party_id: int, year: int, year_work_orders: list) -> list[dict]:
    """12 aylık toplam makbuz verilerini hesapla."""
    monthly_totals = []
    for m in range(1, 13):
        month_wos = [wo for wo in year_work_orders if wo.work_date.month == m]
        month_total = sum((wo.total_price for wo in month_wos), Decimal("0.00"))
        month_makbuz = db.session.execute(
            db.select(Makbuz).where(
                Makbuz.party_id == party_id,
                Makbuz.year == year,
                Makbuz.month == m,
            )
        ).scalar_one_or_none()
        monthly_totals.append({
            "month_number": m,
            "month_name": MONTHS[m - 1][1],
            "count": len(month_wos),
            "total": month_total,
            "vat": month_makbuz.vat_amount if month_makbuz else Decimal("0.00"),
            "grand_total": (month_total + month_makbuz.vat_amount) if month_makbuz else month_total,
            "makbuz_status": month_makbuz.status if month_makbuz else None,
        })
    return monthly_totals


def _compute_previous_debt(party_id: int, year: int, month: int | None, view: str) -> Decimal:
    """Bir önceki dönemlere ait ödenmemiş makbuz toplamını hesapla."""
    previous_periods_debt = Decimal("0.00")
    if view == "month" and month is not None:
        prev_makbuzlar = db.session.execute(
            db.select(Makbuz).where(
                Makbuz.party_id == party_id,
                (Makbuz.year * 100 + Makbuz.month) < (year * 100 + month),
                Makbuz.status.in_((Makbuz.STATUS_SENT, Makbuz.STATUS_PAID)),
            )
        ).scalars().all()
    else:
        prev_makbuzlar = db.session.execute(
            db.select(Makbuz).where(
                Makbuz.party_id == party_id,
                Makbuz.year < year,
                Makbuz.status.in_((Makbuz.STATUS_SENT, Makbuz.STATUS_PAID)),
            )
        ).scalars().all()
    for m in prev_makbuzlar:
        if m.outstanding_amount > 0:
            previous_periods_debt += m.outstanding_amount
    return previous_periods_debt


@parties_bp.route("/<int:party_id>/edit", methods=["GET", "POST"])
@login_required
@permissions_required("clinical.edit")
def edit_party(party_id):
    party = db.get_or_404(Party, party_id)

    if request.method == "POST":
        new_is_active = request.form.get("is_active") == "on"
        if not new_is_active and party.is_active:
            outstanding = _get_party_outstanding(party)
            if outstanding > Decimal("0.00"):
                flash(f"Bu doktorun ₺{outstanding:,.2f} devreden / açık borcu bulunmaktadır. Borcu kapatılmadan pasife alınamaz.", "danger")
                return redirect(url_for("parties.edit_party", party_id=party.id))

        party.name = normalize_display_name(request.form.get("name", ""))
        party.phone = request.form.get("phone", "").strip() or None
        party.email = request.form.get("email", "").strip() or None
        party.address = request.form.get("address", "").strip() or None
        party.tax_id = request.form.get("tax_id", "").strip() or None
        party.notes = request.form.get("notes", "").strip() or None
        party.is_active = new_is_active
        party.previous_balance = money(request.form.get("previous_balance", "0") or "0")

        db.session.commit()
        flash(f"{party.display_name} güncellendi.", "success")
        return redirect(url_for("parties.detail_party", party_id=party.id))

    return render_template("parties/form.html", party=party)


@parties_bp.route("/<int:party_id>/delete", methods=["POST"])
@login_required
@permissions_required("clinical.delete")
def delete_party(party_id):
    party = db.get_or_404(Party, party_id)
    outstanding = _get_party_outstanding(party)
    if outstanding > Decimal("0.00"):
        flash(f"Bu doktorun ₺{outstanding:,.2f} devreden / açık borcu bulunmaktadır. Borcu kapatılmadan doktor silinemez veya pasifleştirilemez.", "danger")
        return redirect(url_for("parties.detail_party", party_id=party.id))

    party.is_active = False
    db.session.commit()
    flash(f"{party.display_name} pasife alındı.", "warning")
    return redirect(url_for("parties.list_parties"))


@parties_bp.route("/<int:party_id>/work-orders/add", methods=["GET", "POST"])
@login_required
@permissions_required("clinical.edit")
def add_work_order(party_id):
    party = PartyService.get_party_or_404(party_id)

    if request.method == "POST":
        try:
            wo = PartyService.create_work_order(party_id, request.form)
            flash("İş emri başarıyla eklendi.", "success")
            return redirect(url_for(
                "parties.detail_party", party_id=party.id,
                year=wo.work_date.year, month=wo.work_date.month,
            ))
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("parties.add_work_order", party_id=party.id))
        except PermissionError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("parties.detail_party", party_id=party.id))

    today = date.today().isoformat()
    return render_template(
        "parties/work_order_form.html",
        party=party,
        work_order=None,
        today=today,
        ana_islemler_treatments=_get_treatments_by_category(TreatmentCategory.ANA_ISLEMLER),
        ekstra_islemler_treatments=_get_treatments_by_category(TreatmentCategory.EKSTRA_ISLEMLER),
        eur_to_try=_get_current_rate(),
        usd_to_try=get_latest_usd_rate() or 0,
    )


@parties_bp.route("/<int:party_id>/work-orders/<int:wo_id>/edit", methods=["GET", "POST"])
@login_required
@permissions_required("clinical.edit")
def edit_work_order(party_id, wo_id):
    party = PartyService.get_party_or_404(party_id)
    wo = PartyService.get_work_order_or_404(party_id, wo_id)

    if request.method == "POST":
        try:
            wo = PartyService.update_work_order(party_id, wo_id, request.form)
            flash("İş emri güncellendi.", "success")
            if request.form.get("return_to") == "work_orders":
                view = request.form.get("return_view", "day")
                if view == "month":
                    return redirect(url_for(
                        "parties.list_work_orders", view="month",
                        year=request.form.get("return_year", type=int),
                        month=request.form.get("return_month", type=int),
                        search=request.form.get("return_search", ""),
                    ))
                return redirect(url_for(
                    "parties.list_work_orders", view="day",
                    date=request.form.get("return_date", ""),
                    search=request.form.get("return_search", ""),
                ))
            if request.form.get("return_to") == "party_detail":
                return redirect(url_for(
                    "parties.detail_party", party_id=party.id,
                    view=request.form.get("return_view", "month"),
                    year=request.form.get("return_year", type=int),
                    month=request.form.get("return_month", type=int),
                    search=request.form.get("return_search", ""),
                ))
            return redirect(url_for("parties.detail_party", party_id=party.id))
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("parties.edit_work_order", party_id=party.id, wo_id=wo.id))
        except PermissionError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("parties.detail_party", party_id=party.id))

    return render_template(
        "parties/work_order_form.html",
        party=party,
        work_order=wo,
        today=wo.work_date.isoformat(),
        ana_islemler_treatments=_get_treatments_by_category(TreatmentCategory.ANA_ISLEMLER),
        ekstra_islemler_treatments=_get_treatments_by_category(TreatmentCategory.EKSTRA_ISLEMLER),
        eur_to_try=_get_current_rate(),
        usd_to_try=get_latest_usd_rate() or 0,
    )


@parties_bp.route("/<int:party_id>/work-orders/<int:wo_id>/delete", methods=["POST"])
@login_required
@permissions_required("clinical.delete")
def delete_work_order(party_id, wo_id):
    try:
        PartyService.delete_work_order(party_id, wo_id)
        flash("İş emri silindi.", "warning")
    except PermissionError as exc:
        flash(str(exc), "danger")

    if request.form.get("return_to") == "work_orders":
        view = request.form.get("return_view", "day")
        if view == "month":
            return redirect(url_for(
                "parties.list_work_orders", view="month",
                year=request.form.get("return_year", type=int),
                month=request.form.get("return_month", type=int),
                search=request.form.get("return_search", ""),
            ))
        return redirect(url_for(
            "parties.list_work_orders", view="day",
            date=request.form.get("return_date", ""),
            search=request.form.get("return_search", ""),
        ))
    return redirect(url_for("parties.detail_party", party_id=party_id))
